import json
import os
import dash
from dash import dcc, html, dash_table
from dash.dependencies import Output, Input, State
import plotly.graph_objs as go
import paho.mqtt.client as mqtt
from collections import deque
from datetime import datetime
import pandas as pd

# Optional dependency for Excel writing
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

# ====== MQTT CONFIG ======
MQTT_BROKER = "broker.hivemq.com"
MQTT_TOPIC = "edukit/suhu"

# ====== DATA BUFFER (3 SENSOR) ======
max_len = 100
timestamps = deque(maxlen=max_len)

# Buffer untuk sensor Air Dingin
data_dingin_c = deque(maxlen=max_len)
data_dingin_f = deque(maxlen=max_len)
data_dingin_k = deque(maxlen=max_len)
data_dingin_r = deque(maxlen=max_len)

# Buffer untuk sensor Air Panas
data_panas_c = deque(maxlen=max_len)
data_panas_f = deque(maxlen=max_len)
data_panas_k = deque(maxlen=max_len)
data_panas_r = deque(maxlen=max_len)

# Buffer untuk sensor Air Campuran
data_campuran_c = deque(maxlen=max_len)
data_campuran_f = deque(maxlen=max_len)
data_campuran_k = deque(maxlen=max_len)
data_campuran_r = deque(maxlen=max_len)

# ====== WARNA UNTUK GRAFIK ======
COLOR_DINGIN = '#1E90FF'   # Biru - Air Dingin
COLOR_PANAS = '#FF4500'    # Merah - Air Panas
COLOR_CAMPURAN = '#32CD32' # Hijau - Air Campuran

# ====== STATUS PENCAMPURAN ======
# Buffer untuk menyimpan nilai kalor secara permanen (tidak dihitung ulang)
kalor_lepas_buffer = deque(maxlen=max_len)   # Nilai Q lepas per data point
kalor_terima_buffer = deque(maxlen=max_len)  # Nilai Q terima per data point

# State pencampuran global (untuk diakses di MQTT callback)
mixing_state_global = {'is_mixing': False, 'massa_dingin': 1.0, 'massa_panas': 1.0}
lock_state_global = {'is_locked': False, 'locked_dingin': 0.0, 'locked_panas': 0.0}

# ====== KALOR CONFIG ======
C_AIR = 4200  # Kalor jenis air dalam J/kg°C
RHO_AIR_DINGIN = 1000  # kg/m^3
RHO_AIR_PANAS = 480    # kg/m^3

# ====== STORAGE CONFIG ======
EXCEL_FILE = "data_suhu.xlsx"
EXCEL_HEADERS = ["Waktu", "Dingin_C", "Dingin_F", "Dingin_K", "Dingin_R", 
                 "Panas_C", "Panas_F", "Panas_K", "Panas_R",
                 "Campuran_C", "Campuran_F", "Campuran_K", "Campuran_R"]

# ====== STORAGE HELPERS ======
def init_excel():
    """Create Excel file with headers if not exists."""
    if Workbook is None or load_workbook is None:
        print("[Excel] openpyxl tidak tersedia. Lewati penyimpanan Excel. Install: pip install openpyxl")
        return
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "DataSuhu"
            ws.append(EXCEL_HEADERS)
            wb.save(EXCEL_FILE)
            print(f"[Excel] File baru dibuat: {EXCEL_FILE}")
    except Exception as e:
        print("[Excel] Gagal inisialisasi file:", e)

def append_row_to_excel(row):
    """Append one row to Excel file. Row example: [timestamp, C, F, K, R]."""
    if Workbook is None or load_workbook is None:
        return
    try:
        if not os.path.exists(EXCEL_FILE):
            init_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row)
        wb.save(EXCEL_FILE)
    except Exception as e:
        print("[Excel] Gagal menyimpan baris:", e)

# ====== MQTT CALLBACK ======
def on_message(client, userdata, msg):
    try:
        payload = json.loads(msg.payload.decode())
        # Pastikan semua data sensor ada
        if all(key in payload for key in ("dingin", "panas", "campuran")):
            dingin = payload["dingin"]
            panas = payload["panas"]
            campuran = payload["campuran"]
            
            # Validasi data setiap sensor
            if all(key in dingin for key in ("C", "F", "K", "R")) and \
               all(key in panas for key in ("C", "F", "K", "R")) and \
               all(key in campuran for key in ("C", "F", "K", "R")):
                
                ts_display = datetime.now().strftime("%H:%M:%S")
                timestamps.append(ts_display)
                
                # Simpan data Air Dingin
                # Cek status lock
                if lock_state_global['is_locked']:
                    val_dingin_c = lock_state_global['locked_dingin']
                    val_panas_c = lock_state_global['locked_panas']
                    # Untuk satuan lain, kita bisa hitung manual atau biarkan (karena yang krusial C)
                    # Untuk simplifikasi visual grafik, kita append nilai locked ke C, 
                    # tapi F, K, R mungkin akan tetap data sensor asli jika tidak kita konversi juga.
                    # Agar konsisten "garis lurus", sebaiknya kita konversi juga atau pakai nilai terakhir.
                    # Namun karena user fokus ke C dan Kalor, kita prioritaskan C.
                    # Biar rapi, kita pakai data sensor asli untuk F, K, R atau biarkan apa adanya?
                    # User minta "grafik ... tetap lurus". Jadi semua satuan harus lurus.
                    # Kita hitung konversi sederhana untuk locked value
                    val_dingin_f = (val_dingin_c * 9/5) + 32
                    val_dingin_k = val_dingin_c + 273.15
                    val_dingin_r = val_dingin_c * 4/5
                    
                    val_panas_f = (val_panas_c * 9/5) + 32
                    val_panas_k = val_panas_c + 273.15
                    val_panas_r = val_panas_c * 4/5
                else:
                    val_dingin_c = dingin["C"]
                    val_dingin_f = dingin["F"]
                    val_dingin_k = dingin["K"]
                    val_dingin_r = dingin["R"]
                    
                    val_panas_c = panas["C"]
                    val_panas_f = panas["F"]
                    val_panas_k = panas["K"]
                    val_panas_r = panas["R"]

                data_dingin_c.append(val_dingin_c)
                data_dingin_f.append(val_dingin_f)
                data_dingin_k.append(val_dingin_k)
                data_dingin_r.append(val_dingin_r)
                
                # Simpan data Air Panas
                data_panas_c.append(val_panas_c)
                data_panas_f.append(val_panas_f)
                data_panas_k.append(val_panas_k)
                data_panas_r.append(val_panas_r)
                
                # Simpan data Air Campuran
                # Cek apakah status finished (freeze result)
                is_finished_global = mixing_state_global.get('is_finished', False)
                
                if is_finished_global:
                    # Gunakan nilai final yang disimpan
                    val_campuran_c = mixing_state_global.get('final_campuran', 0)
                    val_campuran_f = (val_campuran_c * 9/5) + 32
                    val_campuran_k = val_campuran_c + 273.15
                    val_campuran_r = val_campuran_c * 4/5
                else:
                    # Gunakan data real-time
                    val_campuran_c = campuran["C"]
                    val_campuran_f = campuran["F"]
                    val_campuran_k = campuran["K"]
                    val_campuran_r = campuran["R"]
                
                data_campuran_c.append(val_campuran_c)
                data_campuran_f.append(val_campuran_f)
                data_campuran_k.append(val_campuran_k)
                data_campuran_r.append(val_campuran_r)
                
                # Hitung dan simpan nilai kalor berdasarkan mode saat ini
                if mixing_state_global['is_mixing']:
                    # Mode pencampuran - hitung kalor dan simpan secara permanen
                    m_dingin = mixing_state_global['massa_dingin']
                    m_panas = mixing_state_global['massa_panas']
                    # Gunakan nilai yang (mungkin) sudah di-lock
                    q_lepas = abs(m_panas * C_AIR * (val_panas_c - val_campuran_c))
                    q_terima = abs(m_dingin * C_AIR * (val_campuran_c - val_dingin_c))
                    kalor_lepas_buffer.append(q_lepas)
                    kalor_terima_buffer.append(q_terima)
                elif is_finished_global:
                    # Jika finished, kita tetap append nilai kalor terakhir (atau 0?)
                    # Biasanya user ingin melihat nilai kalor terakhir yang "valid"
                    # Ambil nilai terakhir dari buffer jika ada
                    last_q_lepas = kalor_lepas_buffer[-1] if len(kalor_lepas_buffer) > 0 else 0.0
                    last_q_terima = kalor_terima_buffer[-1] if len(kalor_terima_buffer) > 0 else 0.0
                    kalor_lepas_buffer.append(last_q_lepas)
                    kalor_terima_buffer.append(last_q_terima)
                else:
                    # Mode pengukuran awal - simpan 0
                    kalor_lepas_buffer.append(0.0)
                    kalor_terima_buffer.append(0.0)
                
                # Simpan ke Excel
                ts_save = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                append_row_to_excel([
                    ts_save, 
                    val_dingin_c, val_dingin_f, val_dingin_k, val_dingin_r,
                    val_panas_c, val_panas_f, val_panas_k, val_panas_r,
                    campuran["C"], campuran["F"], campuran["K"], campuran["R"]
                ])
                print(f"[MQTT] Data diterima: Dingin={val_dingin_c:.2f}°C, Panas={val_panas_c:.2f}°C, Campuran={campuran['C']}°C")
    except Exception as e:
        print("Gagal parsing data:", e)

# ====== MQTT CLIENT ======
mqtt_client = mqtt.Client()
mqtt_client.on_message = on_message

try:
    # Pastikan file Excel siap
    init_excel()
    mqtt_client.connect(MQTT_BROKER, 1883, 60)
    mqtt_client.subscribe(MQTT_TOPIC)
    mqtt_client.loop_start()
    print(f"Terhubung ke MQTT Broker: {MQTT_BROKER}, Topic: {MQTT_TOPIC}")
except Exception as e:
    print("Gagal konek MQTT:", e)

# ====== DASH APP ======
app = dash.Dash(__name__)
app.layout = html.Div([
    html.H2("🌡️ BlackSense Smart Thermo EduKit Dashboard", style={'textAlign': 'center', 'marginBottom': '5px'}),
    html.H5("Asas Black Learning - Real-Time Heat Transfer Monitoring (3 Sensor)", style={'textAlign': 'center', 'color': '#666', 'fontWeight': 'normal', 'marginTop': '0', 'marginBottom': '20px'}),
    html.Div(id='status', style={'textAlign': 'center', 'color': 'gray', 'marginBottom': '20px'}),
    
    # ====== TOMBOL PENCAMPURAN & STATUS BADGE ======
    html.Div([
        # Tombol Lock Sensor
        html.Button(
            id='btn-lock-sensors',
            children='🔒 Kunci Suhu Awal',
            n_clicks=0,
            style={
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#17a2b8', # Info color
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            }
        ),
        # Tombol Toggle Pencampuran
        html.Button(
            id='btn-toggle-mixing',
            children='🔄 Mulai Pencampuran',
            n_clicks=0,
            style={
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#28a745',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            }
        ),
        # Status Badge
        html.Span(
            id='mixing-status-badge',
            children='Mode: Pengukuran Awal',
            style={
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#6c757d',
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block'
            }
        ),
        # Store untuk menyimpan state pencampuran
        dcc.Store(id='mixing-state', data={'is_mixing': False, 'start_index': -1}),
        # Store untuk menyimpan state lock sensor
        dcc.Store(id='lock-state', data={'is_locked': False, 'locked_temp_dingin': 0, 'locked_temp_panas': 0}),
    ], style={'textAlign': 'center', 'marginBottom': '20px'}),
    
    # Legend Sensor
    html.Div([
        html.Span("● Air Dingin", style={'color': COLOR_DINGIN, 'marginRight': '30px', 'fontWeight': 'bold'}),
        html.Span("● Air Panas", style={'color': COLOR_PANAS, 'marginRight': '30px', 'fontWeight': 'bold'}),
        html.Span("● Air Campuran", style={'color': COLOR_CAMPURAN, 'fontWeight': 'bold'}),
    ], style={'textAlign': 'center', 'marginBottom': '20px', 'fontSize': '16px'}),
    
    # Kontrol dan Card Section
    html.Div([
        # Input Volume untuk masing-masing air
        html.Div([
            html.Div([
                html.Label("Volume Air Dingin (mL):", style={'color': COLOR_DINGIN}),
                dcc.Input(
                    id='volume-dingin-input',
                    type='number',
                    value=250,
                    min=0,
                    step='any',
                    style={'marginLeft': '10px', 'width': '80px'}
                ),
            ], style={'display': 'inline-block', 'marginRight': '30px'}),
            
            html.Div([
                html.Label("Volume Air Panas (mL):", style={'color': COLOR_PANAS}),
                dcc.Input(
                    id='volume-panas-input',
                    type='number',
                    value=250,
                    min=0,
                    step='any',
                    style={'marginLeft': '10px', 'width': '80px'}
                ),
            ], style={'display': 'inline-block', 'marginRight': '30px'}),
        ], style={'marginBottom': '20px'}),

        # Card Kalor - untuk menampilkan kalor yang dipindahkan
        html.Div([
            html.Div([
                html.H4("Kalor Dilepas Air Panas", style={'textAlign': 'center', 'color': COLOR_PANAS}),
                html.Div(id='kalor-dilepas-output', style={'fontSize': '24px', 'textAlign': 'center', 'fontWeight': 'bold', 'color': COLOR_PANAS})
            ], style={'border': f'2px solid {COLOR_PANAS}', 'padding': '20px', 'width': '30%', 'display': 'inline-block', 'margin': '10px', 'borderRadius': '10px'}),
            
            html.Div([
                html.H4("Kalor Diterima Air Dingin", style={'textAlign': 'center', 'color': COLOR_DINGIN}),
                html.Div(id='kalor-diterima-output', style={'fontSize': '24px', 'textAlign': 'center', 'fontWeight': 'bold', 'color': COLOR_DINGIN})
            ], style={'border': f'2px solid {COLOR_DINGIN}', 'padding': '20px', 'width': '30%', 'display': 'inline-block', 'margin': '10px', 'borderRadius': '10px'}),
            
            html.Div([
                html.H4("Suhu Keseimbangan", style={'textAlign': 'center', 'color': COLOR_CAMPURAN}),
                html.Div(id='suhu-campuran-output', style={'fontSize': '24px', 'textAlign': 'center', 'fontWeight': 'bold', 'color': COLOR_CAMPURAN})
            ], style={'border': f'2px solid {COLOR_CAMPURAN}', 'padding': '20px', 'width': '30%', 'display': 'inline-block', 'margin': '10px', 'borderRadius': '10px'})
        ])
    ], style={'textAlign': 'center', 'marginBottom': '30px', 'border': '1px solid #eee', 'padding': '20px', 'width': '90%', 'margin': 'auto', 'borderRadius': '10px'}),
    
    # Grafik Suhu - 2 kolom
    html.Div([
        html.Div([
            dcc.Graph(id='graph-celsius')
        ], style={'width': '48%', 'display': 'inline-block', 'padding': '10px'}),
        
        html.Div([
            dcc.Graph(id='graph-fahrenheit')
        ], style={'width': '48%', 'display': 'inline-block', 'padding': '10px'})
    ]),
    
    html.Div([
        html.Div([
            dcc.Graph(id='graph-kelvin')
        ], style={'width': '48%', 'display': 'inline-block', 'padding': '10px'}),
        
        html.Div([
            dcc.Graph(id='graph-reamur')
        ], style={'width': '48%', 'display': 'inline-block', 'padding': '10px'})
    ]),
    
    html.H4("Tabel Data Real-Time (3 Sensor)", style={'textAlign': 'center', 'marginTop': '40px'}),
    dash_table.DataTable(
        id='live-table',
        columns=[
            # Kolom Waktu
            {'name': ['', 'Waktu'], 'id': 'waktu'},
            # Kolom Air Dingin (4 satuan)
            {'name': ['Air Dingin', '°C'], 'id': 'dingin_c'},
            {'name': ['Air Dingin', '°F'], 'id': 'dingin_f'},
            {'name': ['Air Dingin', 'K'], 'id': 'dingin_k'},
            {'name': ['Air Dingin', '°R'], 'id': 'dingin_r'},
            # Kolom Air Panas (4 satuan)
            {'name': ['Air Panas', '°C'], 'id': 'panas_c'},
            {'name': ['Air Panas', '°F'], 'id': 'panas_f'},
            {'name': ['Air Panas', 'K'], 'id': 'panas_k'},
            {'name': ['Air Panas', '°R'], 'id': 'panas_r'},
            # Kolom Air Campuran (4 satuan)
            {'name': ['Air Campuran', '°C'], 'id': 'campuran_c'},
            {'name': ['Air Campuran', '°F'], 'id': 'campuran_f'},
            {'name': ['Air Campuran', 'K'], 'id': 'campuran_k'},
            {'name': ['Air Campuran', '°R'], 'id': 'campuran_r'},
            # Kolom Kalor
            {'name': ['Kalor', 'Q Lepas (J)'], 'id': 'kalor_lepas'},
            {'name': ['Kalor', 'Q Terima (J)'], 'id': 'kalor_terima'},
        ],
        merge_duplicate_headers=True,
        page_size=15,
        style_cell={
            'textAlign': 'center', 
            'padding': '8px',
            'minWidth': '60px',
            'maxWidth': '100px',
            'whiteSpace': 'normal'
        },
        style_header={
            'backgroundColor': '#f8f9fa',
            'fontWeight': 'bold',
            'border': '1px solid #dee2e6',
            'textAlign': 'center'
        },
        style_data_conditional=[
            # Alternating row colors
            {
                'if': {'row_index': 'odd'},
                'backgroundColor': 'rgb(248, 248, 248)'
            },
            # Air Dingin columns - Biru
            {'if': {'column_id': 'dingin_c'}, 'color': COLOR_DINGIN, 'fontWeight': 'bold'},
            {'if': {'column_id': 'dingin_f'}, 'color': COLOR_DINGIN},
            {'if': {'column_id': 'dingin_k'}, 'color': COLOR_DINGIN},
            {'if': {'column_id': 'dingin_r'}, 'color': COLOR_DINGIN},
            # Air Panas columns - Merah
            {'if': {'column_id': 'panas_c'}, 'color': COLOR_PANAS, 'fontWeight': 'bold'},
            {'if': {'column_id': 'panas_f'}, 'color': COLOR_PANAS},
            {'if': {'column_id': 'panas_k'}, 'color': COLOR_PANAS},
            {'if': {'column_id': 'panas_r'}, 'color': COLOR_PANAS},
            # Air Campuran columns - Hijau
            {'if': {'column_id': 'campuran_c'}, 'color': COLOR_CAMPURAN, 'fontWeight': 'bold'},
            {'if': {'column_id': 'campuran_f'}, 'color': COLOR_CAMPURAN},
            {'if': {'column_id': 'campuran_k'}, 'color': COLOR_CAMPURAN},
            {'if': {'column_id': 'campuran_r'}, 'color': COLOR_CAMPURAN},
        ],
        style_header_conditional=[
            # Header Air Dingin - background biru muda
            {'if': {'column_id': ['dingin_c', 'dingin_f', 'dingin_k', 'dingin_r'], 'header_index': 0},
             'backgroundColor': '#E6F3FF', 'color': COLOR_DINGIN},
            {'if': {'column_id': ['dingin_c', 'dingin_f', 'dingin_k', 'dingin_r'], 'header_index': 1},
             'backgroundColor': '#E6F3FF', 'color': COLOR_DINGIN},
            # Header Air Panas - background merah muda
            {'if': {'column_id': ['panas_c', 'panas_f', 'panas_k', 'panas_r'], 'header_index': 0},
             'backgroundColor': '#FFE6E0', 'color': COLOR_PANAS},
            {'if': {'column_id': ['panas_c', 'panas_f', 'panas_k', 'panas_r'], 'header_index': 1},
             'backgroundColor': '#FFE6E0', 'color': COLOR_PANAS},
            # Header Air Campuran - background hijau muda
            {'if': {'column_id': ['campuran_c', 'campuran_f', 'campuran_k', 'campuran_r'], 'header_index': 0},
             'backgroundColor': '#E6FFE6', 'color': COLOR_CAMPURAN},
            {'if': {'column_id': ['campuran_c', 'campuran_f', 'campuran_k', 'campuran_r'], 'header_index': 1},
             'backgroundColor': '#E6FFE6', 'color': COLOR_CAMPURAN},
            # Header Kalor - background kuning muda
            {'if': {'column_id': ['kalor_lepas', 'kalor_terima'], 'header_index': 0},
             'backgroundColor': '#FFF9E6', 'color': '#856404'},
            {'if': {'column_id': ['kalor_lepas', 'kalor_terima'], 'header_index': 1},
             'backgroundColor': '#FFF9E6', 'color': '#856404'},
        ],
        style_table={'overflowX': 'auto', 'width': '95%', 'margin': 'auto'}
    ),
    html.Button("Export ke Excel", id="btn-export-excel", style={'marginTop': '10px', 'display': 'block', 'margin': 'auto'}),
    dcc.Download(id="download-excel"),
    
    dcc.Interval(id='update', interval=2000, n_intervals=0)
])

# ====== CALLBACK UNTUK LOCK SENSOR ======
@app.callback(
    [Output('lock-state', 'data'),
     Output('btn-lock-sensors', 'children'),
     Output('btn-lock-sensors', 'style')],
    [Input('btn-lock-sensors', 'n_clicks')],
    [State('lock-state', 'data')]
)
def toggle_lock(n_clicks, current_state):
    if n_clicks == 0:
        return (
            {'is_locked': False, 'locked_temp_dingin': 0, 'locked_temp_panas': 0},
            '🔒 Kunci Suhu Awal',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#17a2b8',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            }
        )
    
    is_locked = current_state.get('is_locked', False)
    
    if not is_locked:
        # Kunci sensor - ambil nilai terakhir dari buffer
        last_dingin = data_dingin_c[-1] if len(data_dingin_c) > 0 else 0
        last_panas = data_panas_c[-1] if len(data_panas_c) > 0 else 0
        
        # Ambil timestamp saat ini untuk referensi tabel
        lock_ts = datetime.now().strftime("%H:%M:%S")
        
        # Update global state
        lock_state_global['is_locked'] = True
        lock_state_global['locked_dingin'] = last_dingin
        lock_state_global['locked_panas'] = last_panas
        
        return (
            {'is_locked': True, 'locked_temp_dingin': last_dingin, 'locked_temp_panas': last_panas, 'lock_timestamp': lock_ts},
            '🔓 Buka Kunci Suhu',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#6c757d', # Grey
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            }
        )
    else:
        # Buka kunci
        # Update global state
        lock_state_global['is_locked'] = False
        lock_state_global['locked_dingin'] = 0.0
        lock_state_global['locked_panas'] = 0.0
        
        return (
            {'is_locked': False, 'locked_temp_dingin': 0, 'locked_temp_panas': 0, 'lock_timestamp': None},
            '🔒 Kunci Suhu Awal',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#17a2b8',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            }
        )

# ====== CALLBACK UNTUK TOGGLE PENCAMPURAN ======
@app.callback(
    [Output('mixing-state', 'data'),
     Output('btn-toggle-mixing', 'children'),
     Output('btn-toggle-mixing', 'style'),
     Output('mixing-status-badge', 'children'),
     Output('mixing-status-badge', 'style')],
    [Input('btn-toggle-mixing', 'n_clicks')],
    [State('mixing-state', 'data')]
)
def toggle_mixing(n_clicks, current_state):
    if n_clicks == 0:
        # Initial state
        return (
            {'is_mixing': False, 'is_finished': False, 'final_campuran': 0},
            '🔄 Mulai Pencampuran',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#28a745',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            },
            '📊 Mode: Pengukuran Awal',
            {
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#6c757d',
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block'
            }
        )
    
    is_mixing = current_state.get('is_mixing', False)
    is_finished = current_state.get('is_finished', False)
    
    if not is_mixing and not is_finished:
        # Tahap 1: Mulai Pencampuran
        mixing_state_global['is_mixing'] = True
        return (
            {'is_mixing': True, 'is_finished': False, 'final_campuran': 0},
            '⏹️ Stop & Kunci Hasil',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#dc3545', # Merah
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            },
            '🔥 Mode: Proses Pencampuran',
            {
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#fd7e14', # Orange
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block',
                'animation': 'pulse 1s infinite'
            }
        )
        
    elif is_mixing and not is_finished:
        # Tahap 2: Selesai & Freeze Hasil
        mixing_state_global['is_mixing'] = False # Stop hitung kalor baru
        
        # Ambil suhu campuran terakhir untuk di-freeze
        last_campuran = data_campuran_c[-1] if len(data_campuran_c) > 0 else 0
        
        # Update global state agar on_message tahu harus freeze
        mixing_state_global['is_finished'] = True
        mixing_state_global['final_campuran'] = last_campuran
        
        return (
            {'is_mixing': False, 'is_finished': True, 'final_campuran': last_campuran},
            '🔄 Reset / Ulangi',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#007bff', # Biru
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            },
            '❄️ Mode: Hasil Terkunci',
            {
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#17a2b8', # Cyan
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block'
            }
        )
        
    else:
        # Tahap 3: Reset ke Awal
        mixing_state_global['is_finished'] = False
        mixing_state_global['final_campuran'] = 0
        
        return (
            {'is_mixing': False, 'is_finished': False, 'final_campuran': 0},
            '🔄 Mulai Pencampuran',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#28a745',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            },
            '📊 Mode: Pengukuran Awal',
            {
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#6c757d',
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block'
            }
        )

@app.callback(
    [Output('graph-celsius', 'figure'),
     Output('graph-fahrenheit', 'figure'),
     Output('graph-kelvin', 'figure'),
     Output('graph-reamur', 'figure'),
     Output('status', 'children'),
     Output('live-table', 'data'),
     Output('kalor-diterima-output', 'children'),
     Output('kalor-dilepas-output', 'children'),
     Output('suhu-campuran-output', 'children')],
    [Input('update', 'n_intervals'),
     Input('volume-dingin-input', 'value'),
     Input('volume-panas-input', 'value'),
     Input('mixing-state', 'data'),
     Input('lock-state', 'data')]
)
def update_graph(n, vol_dingin, vol_panas, mixing_state, lock_state):
    # Hitung massa dari volume
    # m = rho * V (V dalam m^3) -> V_mL / 1,000,000
    massa_dingin = 0
    massa_panas = 0
    
    if vol_dingin is not None and vol_dingin > 0:
        massa_dingin = RHO_AIR_DINGIN * (vol_dingin / 1000000)
        mixing_state_global['massa_dingin'] = massa_dingin
        
    if vol_panas is not None and vol_panas > 0:
        massa_panas = RHO_AIR_PANAS * (vol_panas / 1000000)
        mixing_state_global['massa_panas'] = massa_panas
        
    if len(data_dingin_c) < 2 or massa_dingin <= 0 or massa_panas <= 0:
        empty_fig = go.Figure()
        empty_fig.update_layout(title="Menunggu data...")
        status_msg = "Menunggu data dari ESP32 atau masukkan nilai volume yang valid (>0)..."
        kalor_msg = "0 J"
        suhu_msg = "--- °C"
        return empty_fig, empty_fig, empty_fig, empty_fig, status_msg, [], kalor_msg, kalor_msg, suhu_msg
    
    # Ambil status pencampuran
    is_mixing = mixing_state.get('is_mixing', False) if mixing_state else False
    is_finished = mixing_state.get('is_finished', False) if mixing_state else False
    final_campuran_c = mixing_state.get('final_campuran', 0) if mixing_state else 0
    
    # Ambil status lock
    is_locked = lock_state.get('is_locked', False) if lock_state else False
    lock_timestamp = lock_state.get('lock_timestamp', None) if lock_state else None
    
    T_dingin = data_dingin_c[-1]
    T_panas = data_panas_c[-1]
    
    # LOGIKA UTAMA UNTUK POIN 5:
    # Jika is_finished (Selesai & Kunci), gunakan nilai final_campuran yang disimpan
    # Namun, untuk grafik, kita gunakan data_campuran_c yang sudah dimodifikasi di on_message
    # (on_message akan append nilai final jika is_finished=True, sehingga history tetap ada)
    if is_finished:
        T_campuran = final_campuran_c
        # Gunakan data dari buffer yang sudah di-handle oleh on_message
        plot_campuran_c = list(data_campuran_c)
        plot_campuran_f = list(data_campuran_f)
        plot_campuran_k = list(data_campuran_k)
        plot_campuran_r = list(data_campuran_r)
    else:
        # Jika belum selesai, gunakan data real-time
        T_campuran = data_campuran_c[-1]
        plot_campuran_c = list(data_campuran_c)
        plot_campuran_f = list(data_campuran_f)
        plot_campuran_k = list(data_campuran_k)
        plot_campuran_r = list(data_campuran_r)
    
    # Perhitungan Kalor Asas Black - HANYA jika dalam mode pencampuran
    if is_mixing or is_finished:
        # Q lepas (air panas) = m_panas * c * (T_awal_panas - T_campuran)
        # Q terima (air dingin) = m_dingin * c * (T_campuran - T_awal_dingin)
        # Jika locked, T_awal diambil dari nilai locked (yang sudah di-append ke buffer)
        # Karena kita sudah memodifikasi on_message untuk append nilai locked jika is_locked=True,
        # maka T_dingin dan T_panas (data_dingin_c[-1]) sudah pasti nilai locked tersebut.
        # Jadi rumus ini tetap valid tanpa perubahan.
        Q_lepas = massa_panas * C_AIR * (T_panas - T_campuran)
        Q_terima = massa_dingin * C_AIR * (T_campuran - T_dingin)
        kalor_lepas_str = f"{abs(Q_lepas):.2f} J"
        kalor_terima_str = f"{abs(Q_terima):.2f} J"
    else:
        # Sebelum pencampuran - tampilkan 0
        kalor_lepas_str = "0 J"
        kalor_terima_str = "0 J"
    
    suhu_campuran_str = f"{T_campuran:.2f} °C"

    # ====== GRAFIK CELSIUS (Multi-line) ======
    fig_c = go.Figure()
    fig_c.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_dingin_c), 
        mode='lines+markers', name='Air Dingin',
        line=dict(color=COLOR_DINGIN, width=2),
        marker=dict(size=6)
    ))
    fig_c.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_panas_c), 
        mode='lines+markers', name='Air Panas',
        line=dict(color=COLOR_PANAS, width=2),
        marker=dict(size=6)
    ))
    fig_c.add_trace(go.Scatter(
        x=list(timestamps), y=plot_campuran_c, 
        mode='lines+markers', name='Air Campuran',
        line=dict(color=COLOR_CAMPURAN, width=2),
        marker=dict(size=6)
    ))
    fig_c.update_layout(
        title="Suhu Celsius (°C)",
        xaxis_title="Waktu",
        yaxis_title="°C",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # ====== GRAFIK FAHRENHEIT (Multi-line) ======
    fig_f = go.Figure()
    fig_f.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_dingin_f), 
        mode='lines+markers', name='Air Dingin',
        line=dict(color=COLOR_DINGIN, width=2),
        marker=dict(size=6)
    ))
    fig_f.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_panas_f), 
        mode='lines+markers', name='Air Panas',
        line=dict(color=COLOR_PANAS, width=2),
        marker=dict(size=6)
    ))
    fig_f.add_trace(go.Scatter(
        x=list(timestamps), y=plot_campuran_f, 
        mode='lines+markers', name='Air Campuran',
        line=dict(color=COLOR_CAMPURAN, width=2),
        marker=dict(size=6)
    ))
    fig_f.update_layout(
        title="Suhu Fahrenheit (°F)",
        xaxis_title="Waktu",
        yaxis_title="°F",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # ====== GRAFIK KELVIN (Multi-line) ======
    fig_k = go.Figure()
    fig_k.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_dingin_k), 
        mode='lines+markers', name='Air Dingin',
        line=dict(color=COLOR_DINGIN, width=2),
        marker=dict(size=6)
    ))
    fig_k.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_panas_k), 
        mode='lines+markers', name='Air Panas',
        line=dict(color=COLOR_PANAS, width=2),
        marker=dict(size=6)
    ))
    fig_k.add_trace(go.Scatter(
        x=list(timestamps), y=plot_campuran_k, 
        mode='lines+markers', name='Air Campuran',
        line=dict(color=COLOR_CAMPURAN, width=2),
        marker=dict(size=6)
    ))
    fig_k.update_layout(
        title="Suhu Kelvin (K)",
        xaxis_title="Waktu",
        yaxis_title="K",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # ====== GRAFIK REAMUR (Multi-line) ======
    fig_r = go.Figure()
    fig_r.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_dingin_r), 
        mode='lines+markers', name='Air Dingin',
        line=dict(color=COLOR_DINGIN, width=2),
        marker=dict(size=6)
    ))
    fig_r.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_panas_r), 
        mode='lines+markers', name='Air Panas',
        line=dict(color=COLOR_PANAS, width=2),
        marker=dict(size=6)
    ))
    fig_r.add_trace(go.Scatter(
        x=list(timestamps), y=plot_campuran_r, 
        mode='lines+markers', name='Air Campuran',
        line=dict(color=COLOR_CAMPURAN, width=2),
        marker=dict(size=6)
    ))
    fig_r.update_layout(
        title="Suhu Reamur (°R)",
        xaxis_title="Waktu",
        yaxis_title="°R",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # --- Gunakan riwayat kalor dari buffer yang sudah tersimpan ---
    # Nilai kalor disimpan permanen saat data diterima, tidak dihitung ulang
    kalor_lepas_history = list(kalor_lepas_buffer)
    kalor_terima_history = list(kalor_terima_buffer)
    
    # Pastikan panjang buffer kalor sama dengan data lainnya
    # (untuk handle kasus jika ada data lama sebelum buffer kalor ditambahkan)
    while len(kalor_lepas_history) < len(data_dingin_c):
        kalor_lepas_history.insert(0, 0.0)
        kalor_terima_history.insert(0, 0.0)

    # Siapkan data untuk tabel, data terbaru di atas
    # Gunakan plot_campuran_* agar nilai yang ditampilkan di tabel juga ikut "beku" saat finish
    table_data = []
    
    # Pastikan semua list memiliki panjang yang sama sebelum zip
    min_len = min(len(timestamps), len(data_dingin_c), len(plot_campuran_c), len(kalor_lepas_history))
    
    data_list = list(zip(
        list(timestamps)[-min_len:], 
        list(data_dingin_c)[-min_len:], list(data_dingin_f)[-min_len:], list(data_dingin_k)[-min_len:], list(data_dingin_r)[-min_len:],
        list(data_panas_c)[-min_len:], list(data_panas_f)[-min_len:], list(data_panas_k)[-min_len:], list(data_panas_r)[-min_len:],
        plot_campuran_c[-min_len:], plot_campuran_f[-min_len:], plot_campuran_k[-min_len:], plot_campuran_r[-min_len:],
        kalor_lepas_history[-min_len:], kalor_terima_history[-min_len:]
    ))
    
    for row in reversed(data_list):
        ts, dc, df, dk, dr, pc, pf, pk, pr, cc, cf, ck, cr, ql, qt = row
        
        # Jika locked, tampilkan "-" untuk kolom Air Dingin dan Panas
        # HANYA jika timestamp row >= lock_timestamp (data setelah dikunci)
        should_mask = False
        if is_locked and lock_timestamp:
            # Bandingkan string waktu "HH:MM:SS"
            # Asumsi dalam satu hari yang sama
            if ts >= lock_timestamp:
                should_mask = True
        
        if should_mask:
            dc_str = "-"
            df_str = "-"
            dk_str = "-"
            dr_str = "-"
            pc_str = "-"
            pf_str = "-"
            pk_str = "-"
            pr_str = "-"
        else:
            dc_str = f"{dc:.2f}"
            df_str = f"{df:.2f}"
            dk_str = f"{dk:.2f}"
            dr_str = f"{dr:.2f}"
            pc_str = f"{pc:.2f}"
            pf_str = f"{pf:.2f}"
            pk_str = f"{pk:.2f}"
            pr_str = f"{pr:.2f}"
            
        table_data.append({
            'waktu': ts,
            'dingin_c': dc_str,
            'dingin_f': df_str,
            'dingin_k': dk_str,
            'dingin_r': dr_str,
            'panas_c': pc_str,
            'panas_f': pf_str,
            'panas_k': pk_str,
            'panas_r': pr_str,
            'campuran_c': f"{cc:.2f}",
            'campuran_f': f"{cf:.2f}",
            'campuran_k': f"{ck:.2f}",
            'campuran_r': f"{cr:.2f}",
            'kalor_lepas': f"{ql:.2f}",
            'kalor_terima': f"{qt:.2f}"
        })
    
    # Status text dengan indikator mode
    if is_finished:
        mode_indicator = "🏁 SELESAI (HASIL DIKUNCI)"
    elif is_mixing:
        mode_indicator = "⚗️ PENCAMPURAN BERLANGSUNG"
    else:
        mode_indicator = "📊 PENGUKURAN AWAL"
        
    lock_indicator = "🔒 SENSOR AWAL TERKUNCI" if is_locked else "🔓 SENSOR AWAL LIVE"
    
    status_text = f"📡 {MQTT_TOPIC} | {mode_indicator} | {lock_indicator} | Terakhir: {timestamps[-1]} | Dingin: {T_dingin:.1f}°C | Panas: {T_panas:.1f}°C | Campuran: {T_campuran:.1f}°C"
    return fig_c, fig_f, fig_k, fig_r, status_text, table_data, kalor_terima_str, kalor_lepas_str, suhu_campuran_str

@app.callback(
    Output("download-excel", "data"),
    Input("btn-export-excel", "n_clicks"),
    State("live-table", "data"),
    prevent_initial_call=True,
)
def export_table_to_excel(n_clicks, table_data):
    if not table_data:
        return
    
    df = pd.DataFrame(table_data)
    # Rename columns for better Excel readability
    df.columns = [
        'Waktu',
        'Dingin °C', 'Dingin °F', 'Dingin K', 'Dingin °R',
        'Panas °C', 'Panas °F', 'Panas K', 'Panas °R',
        'Campuran °C', 'Campuran °F', 'Campuran K', 'Campuran °R',
        'Q Lepas (J)', 'Q Terima (J)'
    ]
    return dcc.send_data_frame(df.to_excel, "data_tabel.xlsx", sheet_name="DataSuhu", index=False)

if __name__ == "__main__":
    print("Menjalankan dashboard di http://127.0.0.1:8050")
    app.run(debug=True)
