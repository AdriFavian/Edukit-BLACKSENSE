import json
import os
import dash
from dash import dcc, html, dash_table
from dash.dependencies import Output, Input, State
import plotly.graph_objs as go
import paho.mqtt.client as mqtt
from collections import deque
from datetime import datetime
import pandas as pd

# Optional dependency for Excel writing
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

# ====== MQTT CONFIG ======
MQTT_BROKER = "broker.hivemq.com"
MQTT_TOPIC = "edukit/suhu"

# ====== DATA BUFFER (3 SENSOR) ======
max_len = 100
timestamps = deque(maxlen=max_len)

# Buffer untuk sensor Air Dingin
data_dingin_c = deque(maxlen=max_len)
data_dingin_f = deque(maxlen=max_len)
data_dingin_k = deque(maxlen=max_len)
data_dingin_r = deque(maxlen=max_len)

# Buffer untuk sensor Air Panas
data_panas_c = deque(maxlen=max_len)
data_panas_f = deque(maxlen=max_len)
data_panas_k = deque(maxlen=max_len)
data_panas_r = deque(maxlen=max_len)

# Buffer untuk sensor Air Campuran
data_campuran_c = deque(maxlen=max_len)
data_campuran_f = deque(maxlen=max_len)
data_campuran_k = deque(maxlen=max_len)
data_campuran_r = deque(maxlen=max_len)

# ====== WARNA UNTUK GRAFIK ======
COLOR_DINGIN = '#1E90FF'   # Biru - Air Dingin
COLOR_PANAS = '#FF4500'    # Merah - Air Panas
COLOR_CAMPURAN = '#32CD32' # Hijau - Air Campuran

# ====== STATUS PENCAMPURAN ======
# Buffer untuk menyimpan nilai kalor secara permanen (tidak dihitung ulang)
kalor_lepas_buffer = deque(maxlen=max_len)   # Nilai Q lepas per data point
kalor_terima_buffer = deque(maxlen=max_len)  # Nilai Q terima per data point

# State pencampuran global (untuk diakses di MQTT callback)
mixing_state_global = {'is_mixing': False, 'massa_dingin': 1.0, 'massa_panas': 1.0}

# ====== KALOR CONFIG ======
C_AIR = 4186  # Kalor jenis air dalam J/kg°C

# ====== STORAGE CONFIG ======
EXCEL_FILE = "data_suhu.xlsx"
EXCEL_HEADERS = ["Waktu", "Dingin_C", "Dingin_F", "Dingin_K", "Dingin_R", 
                 "Panas_C", "Panas_F", "Panas_K", "Panas_R",
                 "Campuran_C", "Campuran_F", "Campuran_K", "Campuran_R"]

# ====== STORAGE HELPERS ======
def init_excel():
    """Create Excel file with headers if not exists."""
    if Workbook is None or load_workbook is None:
        print("[Excel] openpyxl tidak tersedia. Lewati penyimpanan Excel. Install: pip install openpyxl")
        return
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "DataSuhu"
            ws.append(EXCEL_HEADERS)
            wb.save(EXCEL_FILE)
            print(f"[Excel] File baru dibuat: {EXCEL_FILE}")
    except Exception as e:
        print("[Excel] Gagal inisialisasi file:", e)

def append_row_to_excel(row):
    """Append one row to Excel file. Row example: [timestamp, C, F, K, R]."""
    if Workbook is None or load_workbook is None:
        return
    try:
        if not os.path.exists(EXCEL_FILE):
            init_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row)
        wb.save(EXCEL_FILE)
    except Exception as e:
        print("[Excel] Gagal menyimpan baris:", e)

# ====== MQTT CALLBACK ======
def on_message(client, userdata, msg):
    try:
        payload = json.loads(msg.payload.decode())
        # Pastikan semua data sensor ada
        if all(key in payload for key in ("dingin", "panas", "campuran")):
            dingin = payload["dingin"]
            panas = payload["panas"]
            campuran = payload["campuran"]
            
            # Validasi data setiap sensor
            if all(key in dingin for key in ("C", "F", "K", "R")) and \
               all(key in panas for key in ("C", "F", "K", "R")) and \
               all(key in campuran for key in ("C", "F", "K", "R")):
                
                ts_display = datetime.now().strftime("%H:%M:%S")
                timestamps.append(ts_display)
                
                # Simpan data Air Dingin
                data_dingin_c.append(dingin["C"])
                data_dingin_f.append(dingin["F"])
                data_dingin_k.append(dingin["K"])
                data_dingin_r.append(dingin["R"])
                
                # Simpan data Air Panas
                data_panas_c.append(panas["C"])
                data_panas_f.append(panas["F"])
                data_panas_k.append(panas["K"])
                data_panas_r.append(panas["R"])
                
                # Simpan data Air Campuran
                data_campuran_c.append(campuran["C"])
                data_campuran_f.append(campuran["F"])
                data_campuran_k.append(campuran["K"])
                data_campuran_r.append(campuran["R"])
                
                # Hitung dan simpan nilai kalor berdasarkan mode saat ini
                if mixing_state_global['is_mixing']:
                    # Mode pencampuran - hitung kalor dan simpan secara permanen
                    m_dingin = mixing_state_global['massa_dingin']
                    m_panas = mixing_state_global['massa_panas']
                    q_lepas = abs(m_panas * C_AIR * (panas["C"] - campuran["C"]))
                    q_terima = abs(m_dingin * C_AIR * (campuran["C"] - dingin["C"]))
                    kalor_lepas_buffer.append(q_lepas)
                    kalor_terima_buffer.append(q_terima)
                else:
                    # Mode pengukuran awal - simpan 0
                    kalor_lepas_buffer.append(0.0)
                    kalor_terima_buffer.append(0.0)
                
                # Simpan ke Excel
                ts_save = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                append_row_to_excel([
                    ts_save, 
                    dingin["C"], dingin["F"], dingin["K"], dingin["R"],
                    panas["C"], panas["F"], panas["K"], panas["R"],
                    campuran["C"], campuran["F"], campuran["K"], campuran["R"]
                ])
                print(f"[MQTT] Data diterima: Dingin={dingin['C']}°C, Panas={panas['C']}°C, Campuran={campuran['C']}°C")
    except Exception as e:
        print("Gagal parsing data:", e)

# ====== MQTT CLIENT ======
mqtt_client = mqtt.Client()
mqtt_client.on_message = on_message

try:
    # Pastikan file Excel siap
    init_excel()
    mqtt_client.connect(MQTT_BROKER, 1883, 60)
    mqtt_client.subscribe(MQTT_TOPIC)
    mqtt_client.loop_start()
    print(f"Terhubung ke MQTT Broker: {MQTT_BROKER}, Topic: {MQTT_TOPIC}")
except Exception as e:
    print("Gagal konek MQTT:", e)

# ====== DASH APP ======
app = dash.Dash(__name__)
app.layout = html.Div([
    html.H2("🌡️ BlackSense Smart Thermo EduKit Dashboard", style={'textAlign': 'center', 'marginBottom': '5px'}),
    html.H5("Asas Black Learning - Real-Time Heat Transfer Monitoring (3 Sensor)", style={'textAlign': 'center', 'color': '#666', 'fontWeight': 'normal', 'marginTop': '0', 'marginBottom': '20px'}),
    html.Div(id='status', style={'textAlign': 'center', 'color': 'gray', 'marginBottom': '20px'}),
    
    # ====== TOMBOL PENCAMPURAN & STATUS BADGE ======
    html.Div([
        # Tombol Toggle Pencampuran
        html.Button(
            id='btn-toggle-mixing',
            children='🔄 Mulai Pencampuran',
            n_clicks=0,
            style={
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#28a745',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            }
        ),
        # Status Badge
        html.Span(
            id='mixing-status-badge',
            children='Mode: Pengukuran Awal',
            style={
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#6c757d',
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block'
            }
        ),
        # Store untuk menyimpan state pencampuran
        dcc.Store(id='mixing-state', data={'is_mixing': False, 'start_index': -1}),
    ], style={'textAlign': 'center', 'marginBottom': '20px'}),
    
    # Legend Sensor
    html.Div([
        html.Span("● Air Dingin", style={'color': COLOR_DINGIN, 'marginRight': '30px', 'fontWeight': 'bold'}),
        html.Span("● Air Panas", style={'color': COLOR_PANAS, 'marginRight': '30px', 'fontWeight': 'bold'}),
        html.Span("● Air Campuran", style={'color': COLOR_CAMPURAN, 'fontWeight': 'bold'}),
    ], style={'textAlign': 'center', 'marginBottom': '20px', 'fontSize': '16px'}),
    
    # Kontrol dan Card Section
    html.Div([
        # Input Massa untuk masing-masing air
        html.Div([
            html.Div([
                html.Label("Massa Air Dingin (kg):", style={'color': COLOR_DINGIN}),
                dcc.Input(
                    id='massa-dingin-input',
                    type='number',
                    value=1,
                    min=0,
                    step=0.01,
                    style={'marginLeft': '10px', 'width': '80px'}
                ),
            ], style={'display': 'inline-block', 'marginRight': '30px'}),
            
            html.Div([
                html.Label("Massa Air Panas (kg):", style={'color': COLOR_PANAS}),
                dcc.Input(
                    id='massa-panas-input',
                    type='number',
                    value=1,
                    min=0,
                    step=0.01,
                    style={'marginLeft': '10px', 'width': '80px'}
                ),
            ], style={'display': 'inline-block', 'marginRight': '30px'}),
            
            html.Div([
                html.Label("Massa Air Campuran (kg):", style={'color': COLOR_CAMPURAN}),
                dcc.Input(
                    id='massa-campuran-input',
                    type='number',
                    value=2,
                    min=0,
                    step=0.01,
                    style={'marginLeft': '10px', 'width': '80px'}
                ),
            ], style={'display': 'inline-block'}),
        ], style={'marginBottom': '20px'}),

        # Card Kalor - untuk menampilkan kalor yang dipindahkan
        html.Div([
            html.Div([
                html.H4("Kalor Dilepas Air Panas", style={'textAlign': 'center', 'color': COLOR_PANAS}),
                html.Div(id='kalor-dilepas-output', style={'fontSize': '24px', 'textAlign': 'center', 'fontWeight': 'bold', 'color': COLOR_PANAS})
            ], style={'border': f'2px solid {COLOR_PANAS}', 'padding': '20px', 'width': '30%', 'display': 'inline-block', 'margin': '10px', 'borderRadius': '10px'}),
            
            html.Div([
                html.H4("Kalor Diterima Air Dingin", style={'textAlign': 'center', 'color': COLOR_DINGIN}),
                html.Div(id='kalor-diterima-output', style={'fontSize': '24px', 'textAlign': 'center', 'fontWeight': 'bold', 'color': COLOR_DINGIN})
            ], style={'border': f'2px solid {COLOR_DINGIN}', 'padding': '20px', 'width': '30%', 'display': 'inline-block', 'margin': '10px', 'borderRadius': '10px'}),
            
            html.Div([
                html.H4("Suhu Keseimbangan", style={'textAlign': 'center', 'color': COLOR_CAMPURAN}),
                html.Div(id='suhu-campuran-output', style={'fontSize': '24px', 'textAlign': 'center', 'fontWeight': 'bold', 'color': COLOR_CAMPURAN})
            ], style={'border': f'2px solid {COLOR_CAMPURAN}', 'padding': '20px', 'width': '30%', 'display': 'inline-block', 'margin': '10px', 'borderRadius': '10px'})
        ])
    ], style={'textAlign': 'center', 'marginBottom': '30px', 'border': '1px solid #eee', 'padding': '20px', 'width': '90%', 'margin': 'auto', 'borderRadius': '10px'}),
    
    # Grafik Suhu - 2 kolom
    html.Div([
        html.Div([
            dcc.Graph(id='graph-celsius')
        ], style={'width': '48%', 'display': 'inline-block', 'padding': '10px'}),
        
        html.Div([
            dcc.Graph(id='graph-fahrenheit')
        ], style={'width': '48%', 'display': 'inline-block', 'padding': '10px'})
    ]),
    
    html.Div([
        html.Div([
            dcc.Graph(id='graph-kelvin')
        ], style={'width': '48%', 'display': 'inline-block', 'padding': '10px'}),
        
        html.Div([
            dcc.Graph(id='graph-reamur')
        ], style={'width': '48%', 'display': 'inline-block', 'padding': '10px'})
    ]),
    
    html.H4("Tabel Data Real-Time (3 Sensor)", style={'textAlign': 'center', 'marginTop': '40px'}),
    dash_table.DataTable(
        id='live-table',
        columns=[
            # Kolom Waktu
            {'name': ['', 'Waktu'], 'id': 'waktu'},
            # Kolom Air Dingin (4 satuan)
            {'name': ['Air Dingin', '°C'], 'id': 'dingin_c'},
            {'name': ['Air Dingin', '°F'], 'id': 'dingin_f'},
            {'name': ['Air Dingin', 'K'], 'id': 'dingin_k'},
            {'name': ['Air Dingin', '°R'], 'id': 'dingin_r'},
            # Kolom Air Panas (4 satuan)
            {'name': ['Air Panas', '°C'], 'id': 'panas_c'},
            {'name': ['Air Panas', '°F'], 'id': 'panas_f'},
            {'name': ['Air Panas', 'K'], 'id': 'panas_k'},
            {'name': ['Air Panas', '°R'], 'id': 'panas_r'},
            # Kolom Air Campuran (4 satuan)
            {'name': ['Air Campuran', '°C'], 'id': 'campuran_c'},
            {'name': ['Air Campuran', '°F'], 'id': 'campuran_f'},
            {'name': ['Air Campuran', 'K'], 'id': 'campuran_k'},
            {'name': ['Air Campuran', '°R'], 'id': 'campuran_r'},
            # Kolom Kalor
            {'name': ['Kalor', 'Q Lepas (J)'], 'id': 'kalor_lepas'},
            {'name': ['Kalor', 'Q Terima (J)'], 'id': 'kalor_terima'},
        ],
        merge_duplicate_headers=True,
        page_size=15,
        style_cell={
            'textAlign': 'center', 
            'padding': '8px',
            'minWidth': '60px',
            'maxWidth': '100px',
            'whiteSpace': 'normal'
        },
        style_header={
            'backgroundColor': '#f8f9fa',
            'fontWeight': 'bold',
            'border': '1px solid #dee2e6',
            'textAlign': 'center'
        },
        style_data_conditional=[
            # Alternating row colors
            {
                'if': {'row_index': 'odd'},
                'backgroundColor': 'rgb(248, 248, 248)'
            },
            # Air Dingin columns - Biru
            {'if': {'column_id': 'dingin_c'}, 'color': COLOR_DINGIN, 'fontWeight': 'bold'},
            {'if': {'column_id': 'dingin_f'}, 'color': COLOR_DINGIN},
            {'if': {'column_id': 'dingin_k'}, 'color': COLOR_DINGIN},
            {'if': {'column_id': 'dingin_r'}, 'color': COLOR_DINGIN},
            # Air Panas columns - Merah
            {'if': {'column_id': 'panas_c'}, 'color': COLOR_PANAS, 'fontWeight': 'bold'},
            {'if': {'column_id': 'panas_f'}, 'color': COLOR_PANAS},
            {'if': {'column_id': 'panas_k'}, 'color': COLOR_PANAS},
            {'if': {'column_id': 'panas_r'}, 'color': COLOR_PANAS},
            # Air Campuran columns - Hijau
            {'if': {'column_id': 'campuran_c'}, 'color': COLOR_CAMPURAN, 'fontWeight': 'bold'},
            {'if': {'column_id': 'campuran_f'}, 'color': COLOR_CAMPURAN},
            {'if': {'column_id': 'campuran_k'}, 'color': COLOR_CAMPURAN},
            {'if': {'column_id': 'campuran_r'}, 'color': COLOR_CAMPURAN},
        ],
        style_header_conditional=[
            # Header Air Dingin - background biru muda
            {'if': {'column_id': ['dingin_c', 'dingin_f', 'dingin_k', 'dingin_r'], 'header_index': 0},
             'backgroundColor': '#E6F3FF', 'color': COLOR_DINGIN},
            {'if': {'column_id': ['dingin_c', 'dingin_f', 'dingin_k', 'dingin_r'], 'header_index': 1},
             'backgroundColor': '#E6F3FF', 'color': COLOR_DINGIN},
            # Header Air Panas - background merah muda
            {'if': {'column_id': ['panas_c', 'panas_f', 'panas_k', 'panas_r'], 'header_index': 0},
             'backgroundColor': '#FFE6E0', 'color': COLOR_PANAS},
            {'if': {'column_id': ['panas_c', 'panas_f', 'panas_k', 'panas_r'], 'header_index': 1},
             'backgroundColor': '#FFE6E0', 'color': COLOR_PANAS},
            # Header Air Campuran - background hijau muda
            {'if': {'column_id': ['campuran_c', 'campuran_f', 'campuran_k', 'campuran_r'], 'header_index': 0},
             'backgroundColor': '#E6FFE6', 'color': COLOR_CAMPURAN},
            {'if': {'column_id': ['campuran_c', 'campuran_f', 'campuran_k', 'campuran_r'], 'header_index': 1},
             'backgroundColor': '#E6FFE6', 'color': COLOR_CAMPURAN},
            # Header Kalor - background kuning muda
            {'if': {'column_id': ['kalor_lepas', 'kalor_terima'], 'header_index': 0},
             'backgroundColor': '#FFF9E6', 'color': '#856404'},
            {'if': {'column_id': ['kalor_lepas', 'kalor_terima'], 'header_index': 1},
             'backgroundColor': '#FFF9E6', 'color': '#856404'},
        ],
        style_table={'overflowX': 'auto', 'width': '95%', 'margin': 'auto'}
    ),
    html.Button("Export ke Excel", id="btn-export-excel", style={'marginTop': '10px', 'display': 'block', 'margin': 'auto'}),
    dcc.Download(id="download-excel"),
    
    dcc.Interval(id='update', interval=2000, n_intervals=0)
])

# ====== CALLBACK UNTUK TOGGLE PENCAMPURAN ======
@app.callback(
    [Output('mixing-state', 'data'),
     Output('btn-toggle-mixing', 'children'),
     Output('btn-toggle-mixing', 'style'),
     Output('mixing-status-badge', 'children'),
     Output('mixing-status-badge', 'style')],
    [Input('btn-toggle-mixing', 'n_clicks')],
    [State('mixing-state', 'data')]
)
def toggle_mixing(n_clicks, current_state):
    if n_clicks == 0:
        # Initial state
        return (
            {'is_mixing': False, 'start_index': -1},
            '🔄 Mulai Pencampuran',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#28a745',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            },
            '📊 Mode: Pengukuran Awal',
            {
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#6c757d',
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block'
            }
        )
    
    is_mixing = current_state.get('is_mixing', False)
    
    if not is_mixing:
        # Mulai pencampuran - update global state
        mixing_state_global['is_mixing'] = True
        start_idx = len(data_dingin_c) - 1 if len(data_dingin_c) > 0 else 0
        return (
            {'is_mixing': True, 'start_index': start_idx},
            '⏹️ Selesai Pencampuran',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#dc3545',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            },
            '🔥 Mode: Proses Pencampuran',
            {
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#fd7e14',
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block',
                'animation': 'pulse 1s infinite'
            }
        )
    else:
        # Selesai pencampuran - update global state
        mixing_state_global['is_mixing'] = False
        return (
            {'is_mixing': False, 'start_index': -1},
            '🔄 Mulai Pencampuran',
            {
                'padding': '15px 30px',
                'fontSize': '18px',
                'fontWeight': 'bold',
                'backgroundColor': '#28a745',
                'color': 'white',
                'border': 'none',
                'borderRadius': '10px',
                'cursor': 'pointer',
                'marginRight': '20px'
            },
            '📊 Mode: Pengukuran Awal',
            {
                'padding': '10px 20px',
                'fontSize': '16px',
                'fontWeight': 'bold',
                'backgroundColor': '#6c757d',
                'color': 'white',
                'borderRadius': '20px',
                'display': 'inline-block'
            }
        )

@app.callback(
    [Output('graph-celsius', 'figure'),
     Output('graph-fahrenheit', 'figure'),
     Output('graph-kelvin', 'figure'),
     Output('graph-reamur', 'figure'),
     Output('status', 'children'),
     Output('live-table', 'data'),
     Output('kalor-diterima-output', 'children'),
     Output('kalor-dilepas-output', 'children'),
     Output('suhu-campuran-output', 'children')],
    [Input('update', 'n_intervals'),
     Input('massa-dingin-input', 'value'),
     Input('massa-panas-input', 'value'),
     Input('massa-campuran-input', 'value'),
     Input('mixing-state', 'data')]
)
def update_graph(n, massa_dingin, massa_panas, massa_campuran, mixing_state):
    # Update global state dengan nilai massa terbaru
    if massa_dingin and massa_dingin > 0:
        mixing_state_global['massa_dingin'] = massa_dingin
    if massa_panas and massa_panas > 0:
        mixing_state_global['massa_panas'] = massa_panas
        
    if len(data_dingin_c) < 2 or massa_dingin is None or massa_panas is None or massa_dingin <= 0 or massa_panas <= 0:
        empty_fig = go.Figure()
        empty_fig.update_layout(title="Menunggu data...")
        status_msg = "Menunggu data dari ESP32 atau masukkan nilai massa yang valid (>0)..."
        kalor_msg = "0 J"
        suhu_msg = "--- °C"
        return empty_fig, empty_fig, empty_fig, empty_fig, status_msg, [], kalor_msg, kalor_msg, suhu_msg
    
    # Ambil status pencampuran
    is_mixing = mixing_state.get('is_mixing', False) if mixing_state else False
    start_index = mixing_state.get('start_index', -1) if mixing_state else -1
    
    T_dingin = data_dingin_c[-1]
    T_panas = data_panas_c[-1]
    T_campuran = data_campuran_c[-1]
    
    # Perhitungan Kalor Asas Black - HANYA jika dalam mode pencampuran
    if is_mixing:
        # Q lepas (air panas) = m_panas * c * (T_awal_panas - T_campuran)
        # Q terima (air dingin) = m_dingin * c * (T_campuran - T_awal_dingin)
        Q_lepas = massa_panas * C_AIR * (T_panas - T_campuran)
        Q_terima = massa_dingin * C_AIR * (T_campuran - T_dingin)
        kalor_lepas_str = f"{abs(Q_lepas):.2f} J"
        kalor_terima_str = f"{abs(Q_terima):.2f} J"
    else:
        # Sebelum pencampuran - tampilkan 0
        kalor_lepas_str = "0 J"
        kalor_terima_str = "0 J"
    
    suhu_campuran_str = f"{T_campuran:.2f} °C"

    # ====== GRAFIK CELSIUS (Multi-line) ======
    fig_c = go.Figure()
    fig_c.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_dingin_c), 
        mode='lines+markers', name='Air Dingin',
        line=dict(color=COLOR_DINGIN, width=2),
        marker=dict(size=6)
    ))
    fig_c.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_panas_c), 
        mode='lines+markers', name='Air Panas',
        line=dict(color=COLOR_PANAS, width=2),
        marker=dict(size=6)
    ))
    fig_c.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_campuran_c), 
        mode='lines+markers', name='Air Campuran',
        line=dict(color=COLOR_CAMPURAN, width=2),
        marker=dict(size=6)
    ))
    fig_c.update_layout(
        title="Suhu Celsius (°C)",
        xaxis_title="Waktu",
        yaxis_title="°C",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # ====== GRAFIK FAHRENHEIT (Multi-line) ======
    fig_f = go.Figure()
    fig_f.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_dingin_f), 
        mode='lines+markers', name='Air Dingin',
        line=dict(color=COLOR_DINGIN, width=2),
        marker=dict(size=6)
    ))
    fig_f.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_panas_f), 
        mode='lines+markers', name='Air Panas',
        line=dict(color=COLOR_PANAS, width=2),
        marker=dict(size=6)
    ))
    fig_f.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_campuran_f), 
        mode='lines+markers', name='Air Campuran',
        line=dict(color=COLOR_CAMPURAN, width=2),
        marker=dict(size=6)
    ))
    fig_f.update_layout(
        title="Suhu Fahrenheit (°F)",
        xaxis_title="Waktu",
        yaxis_title="°F",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # ====== GRAFIK KELVIN (Multi-line) ======
    fig_k = go.Figure()
    fig_k.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_dingin_k), 
        mode='lines+markers', name='Air Dingin',
        line=dict(color=COLOR_DINGIN, width=2),
        marker=dict(size=6)
    ))
    fig_k.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_panas_k), 
        mode='lines+markers', name='Air Panas',
        line=dict(color=COLOR_PANAS, width=2),
        marker=dict(size=6)
    ))
    fig_k.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_campuran_k), 
        mode='lines+markers', name='Air Campuran',
        line=dict(color=COLOR_CAMPURAN, width=2),
        marker=dict(size=6)
    ))
    fig_k.update_layout(
        title="Suhu Kelvin (K)",
        xaxis_title="Waktu",
        yaxis_title="K",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # ====== GRAFIK REAMUR (Multi-line) ======
    fig_r = go.Figure()
    fig_r.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_dingin_r), 
        mode='lines+markers', name='Air Dingin',
        line=dict(color=COLOR_DINGIN, width=2),
        marker=dict(size=6)
    ))
    fig_r.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_panas_r), 
        mode='lines+markers', name='Air Panas',
        line=dict(color=COLOR_PANAS, width=2),
        marker=dict(size=6)
    ))
    fig_r.add_trace(go.Scatter(
        x=list(timestamps), y=list(data_campuran_r), 
        mode='lines+markers', name='Air Campuran',
        line=dict(color=COLOR_CAMPURAN, width=2),
        marker=dict(size=6)
    ))
    fig_r.update_layout(
        title="Suhu Reamur (°R)",
        xaxis_title="Waktu",
        yaxis_title="°R",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # --- Gunakan riwayat kalor dari buffer yang sudah tersimpan ---
    # Nilai kalor disimpan permanen saat data diterima, tidak dihitung ulang
    kalor_lepas_history = list(kalor_lepas_buffer)
    kalor_terima_history = list(kalor_terima_buffer)
    
    # Pastikan panjang buffer kalor sama dengan data lainnya
    # (untuk handle kasus jika ada data lama sebelum buffer kalor ditambahkan)
    while len(kalor_lepas_history) < len(data_dingin_c):
        kalor_lepas_history.insert(0, 0.0)
        kalor_terima_history.insert(0, 0.0)

    # Siapkan data untuk tabel, data terbaru di atas
    table_data = []
    data_list = list(zip(
        timestamps, 
        data_dingin_c, data_dingin_f, data_dingin_k, data_dingin_r,
        data_panas_c, data_panas_f, data_panas_k, data_panas_r,
        data_campuran_c, data_campuran_f, data_campuran_k, data_campuran_r,
        kalor_lepas_history, kalor_terima_history
    ))
    
    for row in reversed(data_list):
        ts, dc, df, dk, dr, pc, pf, pk, pr, cc, cf, ck, cr, ql, qt = row
        table_data.append({
            'waktu': ts,
            'dingin_c': f"{dc:.2f}",
            'dingin_f': f"{df:.2f}",
            'dingin_k': f"{dk:.2f}",
            'dingin_r': f"{dr:.2f}",
            'panas_c': f"{pc:.2f}",
            'panas_f': f"{pf:.2f}",
            'panas_k': f"{pk:.2f}",
            'panas_r': f"{pr:.2f}",
            'campuran_c': f"{cc:.2f}",
            'campuran_f': f"{cf:.2f}",
            'campuran_k': f"{ck:.2f}",
            'campuran_r': f"{cr:.2f}",
            'kalor_lepas': f"{ql:.2f}",
            'kalor_terima': f"{qt:.2f}"
        })
    
    # Status text dengan indikator mode
    mode_indicator = "🔥 PENCAMPURAN" if is_mixing else "📊 PENGUKURAN"
    status_text = f"📡 {MQTT_TOPIC} | {mode_indicator} | Terakhir: {timestamps[-1]} | Dingin: {T_dingin:.1f}°C | Panas: {T_panas:.1f}°C | Campuran: {T_campuran:.1f}°C"
    return fig_c, fig_f, fig_k, fig_r, status_text, table_data, kalor_terima_str, kalor_lepas_str, suhu_campuran_str

@app.callback(
    Output("download-excel", "data"),
    Input("btn-export-excel", "n_clicks"),
    State("live-table", "data"),
    prevent_initial_call=True,
)
def export_table_to_excel(n_clicks, table_data):
    if not table_data:
        return
    
    df = pd.DataFrame(table_data)
    # Rename columns for better Excel readability
    df.columns = [
        'Waktu',
        'Dingin °C', 'Dingin °F', 'Dingin K', 'Dingin °R',
        'Panas °C', 'Panas °F', 'Panas K', 'Panas °R',
        'Campuran °C', 'Campuran °F', 'Campuran K', 'Campuran °R',
        'Q Lepas (J)', 'Q Terima (J)'
    ]
    return dcc.send_data_frame(df.to_excel, "data_tabel.xlsx", sheet_name="DataSuhu", index=False)

if __name__ == "__main__":
    print("Menjalankan dashboard di http://127.0.0.1:8050")
    app.run(debug=True)
